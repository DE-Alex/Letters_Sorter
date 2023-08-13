import os, shutil
import re
import xlsxwriter


import MyLibs.EXCEL_read_and_write as Excel
import MyLibs.Scan_DirsFiles as Scan
import MyLibs.TXT_RW as TXT_RW

format = ['Адресат', 'Вх/Исх', '№', 'Дата', 'Тема', 'Файл с письмом', 'Ответ', 'Содержание ответа', 'Срок', 'Отв. лицо', 'Файл с ответом']
FilesToSearch = ['.pdf']
SkipFiles = ['Thumbs.db', 'РеестрПисем.xlsx']
SkipFileNames = ['Приложение']


def SplitFileName(FilePath, RootFolder):
	newpath = FilePath.lstrip(RootFolder + '\\')
	tmp = newpath.split('\\')
	KAname, InOut, filename = tmp
	InOut = RootFolder + '\\' + KAname + '\\' + InOut
	
	ExtRe = r'(.\w{3,4})$'
	match = re.compile('(.+)' + ExtRe).search(filename)
	tmp, Ext = match.groups()
	
	AttachRe = r'приложени.+'
	match = re.compile(AttachRe, re.IGNORECASE).search(tmp)
	if match == None: Attach = None
	else:
		match = re.compile('(.+)' + r'(Приложени.+)$').search(tmp)
		tmp, Attach = match.groups()
	
	tmp2 = tmp.split(' ')
	if len(tmp2) == 3: Date, Number, Topic = tmp2
	else: 
		DateRe = r'(\d{4}.\d{2}.\d{2})'
		match = re.compile(DateRe + '(.+)').search(tmp)
		Date, tmp = match.groups()
		
		NumRe = r'([0-9-_]+)'
		match = re.compile(NumRe + '(.+)').search(tmp)
		Number, Topic = match.groups()
		
		Number = Number.strip(' ')
		Number = Number.strip('_')
		Topic = Topic.strip(' ')
		Topic = Topic.strip('_')
	print(f'Date:{Date}, Number:{Number}, Topic:{Topic}, Attachment:{Attach}, Ext:{Ext}')
	return KAname, InOut, Number, Date, Topic, FilePath, Attach, Ext

	
def Reestr(path):
	data = ['Дата','Вх./Исх.','Адресат', 'Содержание', 'Гиперссылка']
	Excel.Read_dict_2D_arrays(path)

def ReadXLS(path):
	from openpyxl import load_workbook
	wb = load_workbook(path)
	SheetName = wb.sheetnames[0]
	ws = wb[SheetName]
	Data = []
	LinksToFiles = []
	
	
	for row in ws.iter_rows(min_row=2, values_only = True):
		Data.append(row)
		LinksToFiles.append(row[5])
	return Data, LinksToFiles
	

def WriteXLS(Data):
	Today = datetime_to_str(now_local(), '%Y.%m.%d')
	path = rf'D:\@Письма\{Today}РеестрПисем.xlsx'
	
	#Create new file
	workbook = xlsxwriter.Workbook(path)
	ws_Reestr = workbook.add_worksheet('Реестр')
	ws_Documents = workbook.add_worksheet('Документы')
	ws_Other = workbook.add_worksheet('Другое')
	# Add a bold format to use to highlight cells.
	bold = workbook.add_format({'bold': True})

	#for SheetData in AllSheetsData:
	for row_num, row in enumerate(Data):
		for col_num, data in enumerate(row):
			if type(data) == str:
				if data.startswith('\\\\') or data.startswith('D:\\'):
					if data.endswith('\Исходящие'):	ws_Reestr.write_url(row_num, col_num, data, string = 'Исходящие')
					elif data.endswith('\Входящие'): ws_Reestr.write_url(row_num, col_num, data, string = 'Входящие')
					else: ws_Reestr.write_url(row_num, col_num, data)
				else: ws_Reestr.write(row_num, col_num, data)
			else: ws_Reestr.write(row_num, col_num, data)

	
	workbook.close()

	
if __name__ == '__main__':
	RootFolder = 'D:\@Письма'
	Data, LinksToFiles  = ReadXLS(RootFolder + '\\'+ 'РеестрПисем.xlsx')

	_, DirPaths = Scan.DirScanPaths(RootFolder) #Find all folders in Path. Skip files in root
	print(DirPaths)
	input()
	SubDirs = []
	for dir in DirPaths:
		_, tmp = Scan.DirScanPaths(dir) #Find all subfolders in folder (Исходящие и Входящие). Skip files in folder
		SubDirs.extend(tmp)
	PathToFiles = []
	for dir in SubDirs:
		tmp, _ = Scan.SubdirScanPaths(dir)
		PathToFiles.extend(tmp)

	for FilePath in PathToFiles:
		KAname, InOut, Number, Date, Topic, FilePath, Attach, Ext = SplitFileName(FilePath, RootFolder)
				
		if FilePath in LinksToFiles:
			N = LinksToFiles.index(FilePath)
			print(N)
			row = Data[N]

		elif FilePath.endswith('.pdf') and Attach == None:
			Data.append((KAname, InOut, Number, Date, Topic, FilePath, None, None, None, None, None))
			
	
	WriteXLS(Data)
	
	