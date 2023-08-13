import os, shutil, re, sys
from datetime import datetime
import MyLibs.Scan_DirsFiles as Scan
import MyLibs.PyObject_to_PyFile as PyFile
import MyLibs.Link_to_Filename as Link

pathToScan = r'D:\@Письма'
Reestr = pathToScan + '\\'+ 'РеестрПисем.xlsx'

Today = datetime_to_str(now_local(), '%Y.%m.%d')	
SaveTo = Reestr
logfile = sys.path[0] + rf'\log.txt'

#Folders to scan
FoldersToSkip = '@'
FldIn = 'Входящие'
FldOut = 'Исходящие'
FldDocs = 'Документы'

#Excel Table
ShtReestr = 'Письма'
ShtDocs = 'Документы'
ShtOther = 'Списки'

class Letter():
	def __init__(self): 
		with open(logfile, 'w') as file: pass

	def ReadXLS(self):
		print('Read Table...', end='')
		from openpyxl import load_workbook
		WBook = load_workbook(Reestr)
		self.Data = {}
		links = []
		for name in WBook.sheetnames:
			ws = WBook[name]
			rows = []
			for row in ws.iter_rows():	#read cell as Excel objects with all attributes
				tmp = []
				for cell in row:
					try: 
						hyperlink = cell.hyperlink.target
						relLink = hyperlink.lstrip('file:///')
						relLink = relLink.replace(pathToScan + '\\', '')
						relLink = relLink.replace('%20', ' ')
						relLink = relLink.replace('/', '\\')
						links.append(relLink)
					except: relLink = None
					value = cell.value
					if type(value) == datetime: value = datetime.strftime(value, "%Y.%m.%d")
					if type(value) == int: value = str(value)
					tmp.append([value, relLink])
				rows.append(tmp)
			self.Data[name] = rows
		print('OK')
		linkfile = sys.path[0] + rf'\link.txt'
		PyFile.Write(links, linkfile)
		
	def SortFiles(self):
		print('Scan Files...', end='')
		FilePaths, _ = Scan.SubdirScanPaths(pathToScan)
		self.Letters, self.Attach, self.Other, self.Docs, self.Trash, = [], [], [], [], []
		self.Word, self.ZIP = [], []
		for absPath in FilePaths:
			relPath = absPath.lstrip('file:///')
			relPath = relPath.replace(pathToScan + '\\', '')
			relPath = relPath.replace('%20', ' ')
			
			tmp = relPath.split('\\')
			if FoldersToSkip in tmp[0]: continue #Skip folders with '@'
			elif len(tmp) == 1: continue #skip files in root
			elif len(tmp) == 2: self.Trash.append(relPath) #files to clean in Adressat folders
			elif len(tmp) == 3:
				if (FldIn in tmp[1]) or (FldOut in tmp[1]):
					if re.compile(r'приложени.+', re.IGNORECASE).search(tmp[-1]): self.Attach.append(relPath)#Приложения
					elif re.compile('.pdf', re.IGNORECASE).search(tmp[-1]): self.Letters.append(relPath)
					elif re.compile('.doc', re.IGNORECASE).search(tmp[-1]): self.Word.append(relPath)
					elif re.compile('.zip', re.IGNORECASE).search(tmp[-1]): self.ZIP.append(relPath)
					else: self.Trash.append(relPath)#files to clean in FldIn/FldOut
				elif FldDocs in tmp[1]: self.Docs.append(relPath)
				else: input(f'!. New folder found (with files)({absPath})')
			elif len(tmp) > 3: 
				if (FldIn in tmp[1]) or (FldOut in tmp[1]): self.Trash.append(relPath)#files to clean in FldIn/FldOut
				elif FldDocs in tmp[1]: self.Docs.append(relPath)
			else: input(f'!. Not classified: ({absPath})')
		self.Other = self.Word + self.ZIP + self.Attach
		print('OK')
		print(f'Found:')
		print(f'Letters - {len(self.Letters)}')
		print(f'Attach - {len(self.Attach)}')
		print(f'Word - {len(self.Word)}')
		print(f'ZIP - {len(self.ZIP)}')
		print(f'Other - {len(self.Other)}')
		print(f'Docs - {len(self.Docs)}')
		print(f'Trash - {len(self.Trash)}')
		urlfile = sys.path[0] + rf'\url.txt'
		data = [['Letters\n'] + self.Letters + ['Attach\n'] + self.Attach + ['Word\n'] + self.Word + ['ZIP\n'] + self.ZIP + ['Docs\n'] + self.Docs + ['Trash\n'] + self.Trash]
		PyFile.Write(data, urlfile)
	
	def UpdateLetters(self):
		Col_Adr = 0							#'Адресат' (контрагент)
		Col_InOut, Col_FolderLink = 1, 1	#'Вх/Исх' and hyperlink to folder
		Col_Num, Col_FileLink = 2, 2		#'№' and hyperlink to file
		Col_Date = 3						#'Дата'
		Col_Subj = 4						#'Тема'
			
		rows = self.Data[ShtReestr]
		
		Title = rows[0]
		UpdatedData = []
		Report = []
		UpdatedData.append(rows[0]) #Title
		LinksToFiles = [row[Col_FileLink][1] for row in rows]

		LTF = sys.path[0] + rf'\LTF.txt'
		PyFile.Write(LinksToFiles, LTF)
		
		FilesToRename = []
		SumNew, SumUpdate = 0, 0
		for relPath in self.Letters:
			Adr, Folder, Number, Date, Topic, Ext = self.SplitLetters(relPath)
		
			if relPath in LinksToFiles:
				N = LinksToFiles.index(relPath)
				row = rows[N]
				row_old = row
				if row[Col_Adr][0] == None: row[Col_Adr][0] = Adr
				
				if row[Col_InOut][0] == None:
					if 'Исх' in Folder: row[Col_InOut][0] = 'Исх.'
					if 'Вх' in Folder: row[Col_InOut][0] = 'Вх.'
				elif row[Col_InOut][0] == 'Вх.': Folder = FldIn
				elif row[Col_InOut][0] == 'Исх.': Folder = FldOut
				
				if row[Col_Date][0] == None: row[Col_Date][0] = Date
				if row[Col_Num][0] == None: row[Col_Num][0] = Number.replace('_', '/')
				if row[Col_Subj][0] == None: row[Col_Subj][0] = Topic.replace('_', ' ')
				
				old, new = [], []
				for i, j in zip(row_old, row):
					for a, b in zip(i,j):
						if a != b: 
							old.append(a)
							new.append(b)
				if new != []:
					SumUpdate += 1
					msg = f'UPDATE {SumUpdate}: in {row_old} updated {new}'
					print(msg)
					with open(logfile, 'a') as file: file.write(msg + '\n')

					
				newDate = Link.Clear(row[Col_Date][0])
				newNum = Link.Clear(row[Col_Num][0]).replace(' ', '_')
				newSubj = Link.Clear(row[Col_Subj][0]).replace(' ', '_')
				for i in ['_',' ']:
					newNum = newNum.strip(i)
					newSubj = newSubj.strip(i)
			
				newFilename = newDate + ' ' + newNum + ' ' + newSubj + Ext
				
				newRelPath = row[Col_Adr][0] + '\\' + Folder + '\\' + newFilename
				if relPath != newRelPath: 
					FilesToRename.append((N, relPath, newRelPath))

			else:
				row = [[None, None] for i in range(len(Title))]
				row[Col_Adr][0] = Adr
				if 'Исх' in Folder: row[Col_InOut][0] = 'Исх.'
				if 'Вх' in Folder: row[Col_InOut][0] = 'Вх.'
				row[Col_Date][0] = Date
				row[Col_Num][0] = Number.replace('_', '/')
				row[Col_Subj][0] = Topic.replace('_', ' ')
				
				row[Col_FolderLink][1] = Adr + '\\' + Folder
				row[Col_FileLink][1] = relPath
				
				SumNew += 1
				msg = f'NEW {SumNew}:\n{row}'
				print(msg)
				with open(logfile, 'a') as file: file.write(msg + '\n')
			UpdatedData.append(row)
		print(f'Total {SumUpdate} records updated, {SumNew} records added' + '\n')
		print(f'Files to rename: {len(FilesToRename)}')
		
		Rename = None
		for N, relPath, newRelPath in FilesToRename:
			absPath = pathToScan + '\\' + relPath
			newAbsPath = pathToScan + '\\' + newRelPath
			msg = f'from: {relPath}\n  to: {newRelPath}...'
			print(msg, end='')
			if Rename == 'A': 
				self.MoveFile(absPath, newAbsPath)
				UpdatedData[N][Col_FileLink][1] = newRelPath
				print('Ok')
				with open(logfile, 'a') as file: file.write(msg + '\n')
				self.moveOther(newRelPath)
			else:
				print('\nRename?: yes(y) / no(n) / All(A) / None(N)...',end='')
				Rename = input()
				if Rename == 'N': break
				if Rename == 'n': 
					continue
				if Rename == 'y' or Rename == 'A':
					self.MoveFile(absPath, newAbsPath)
					UpdatedData[N][Col_FileLink][1] = newRelPath
					print('Ok')
					with open(logfile, 'a') as file: file.write(msg + '\n')
					self.moveOther(newRelPath)
		
		print('Update links...', end = '')
		for row in UpdatedData[1:]:
			FileLink = row[Col_FileLink][1]
			Adr, Folder, _, _, _, _ = self.SplitLetters(FileLink)
			row[Col_FolderLink][1] = Adr + '\\' + Folder
		print('Ok')
		
		self.Data[ShtReestr] = UpdatedData
		print('Reestr updated')
				
	def moveOther(self, newRelPath):
		Adr, Folder, Number, Date, Topic, Ext = self.SplitLetters(newRelPath)
		OtherFiles = []
		for relAttPath in self.Other:
			if (Adr in relAttPath) and (Folder in relAttPath) and (Date in relAttPath) and (Number in relAttPath):
				_, _, _, _, _, Ext = self.SplitLetters(relAttPath)
				match = re.compile('(.+)' + '(приложени.+)', re.IGNORECASE).search(relAttPath)
				if match == None:
					newfilename = Date + ' ' + Number + ' ' + Topic + Ext
				else:
					_, Att = match.groups()
					newfilename = Link.Clear(Date + ' ' + Number + ' ' + Att)
					
					
				NewRelAttPath = Adr + '\\' + Folder + '\\' + newfilename
				if relAttPath != NewRelAttPath:
					msg = f'from: {relAttPath}\n  to: {NewRelAttPath}...'
					print(msg)
					self.MoveFile(pathToScan + '\\' + relAttPath, pathToScan + '\\' + NewRelAttPath)
					with open(logfile, 'a') as file: file.write(msg + '\n')
				OtherFiles.append(NewRelAttPath)
			else: OtherFiles.append(relAttPath)
		self.Other = OtherFiles
		
		
	def SplitLetters(self, relPath):
		Adr, Folder, filename = relPath.split('\\')
		
		ExtRe = r'(\.\w{3,4})$'
		match = re.compile('(.+)' + ExtRe).search(filename)
		tmp, Ext = match.groups()

		try:
			tmp2 = tmp.split(' ')
			if len(tmp2) == 3: Date, Number, Topic = tmp2
			else:
				DateRe = r'(\d{4}.\d{2}.\d{2})'
				match = re.compile(DateRe + '(.+)').search(tmp)
				Date, tmp = match.groups()
				
				NumRe = r'([0-9-_]+)'
				match = re.compile(NumRe + '(.+)').search(tmp)
				Number, Topic = match.groups()
			for i in ['_',' ']:
				Number = Number.strip(i)
				Topic = Topic.strip(i)
		except:
			Date, Number, Topic = '-------', tmp, '-------'
		return Adr, Folder, Number, Date, Topic, Ext

	
	def MoveFile(self, oldPath, newPath):
		tmp = newPath.split('\\')
		newdir = ('\\').join(tmp[:-1])
		if not os.path.exists(newdir): os.makedirs(newdir)
		while True:
			if os.path.exists(newPath): 
				print('File already exists!')
				input()
			else: break
		while True:
			try:
				shutil.move(oldPath, newPath)
				break
			except PermissionError as e:
				print(e)
				input()
		
	def WriteXLS(self):
		Col_Width = {
			ShtReestr:{'A:A' : 15, 'B:B' : 6.22, 'C:C' : 13.33, 'D:D' : 9.33, 'E:E' : 59, 'F:F' : 8.11},
			ShtDocs:{'A:A' : 8.11},
			ShtOther:{'A:A' : 8.11}
			}
		Col_Center = [0,1,2,3,5,6]
		print('Format data to XLSX ...', end=' ')
		import xlsxwriter
		#Create new file
		workbook = xlsxwriter.Workbook(SaveTo)
		
		#FORMAT
		default = workbook.add_format({'border':1})
		title_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color':'#D9D9D9', 'border':1})
		center = workbook.add_format({'align': 'center', 'border':1})
		URL_center = workbook.add_format({'align': 'center', 'font_color': '#0000FF', 'underline': '1', 'border':1})
		
		for ShtName, rows in self.Data.items(): 
			ws = workbook.add_worksheet(ShtName)
			for col, width in Col_Width[ShtName].items(): ws.set_column(col, width)
			
			#Title
			if rows != []:
				title = [value for (value, hLink) in rows[0]]
				ws.write_row ('A1', title, title_format)
			for row_num, row in enumerate(rows[1:]):
				for col_num, data in enumerate(row):
					if col_num in Col_Center: cell_format = center
					else: cell_format = default
					value, hLink = data
					if hLink == None: 
						if type(value) == 'datetime.datetime': 
							print(value)
							ws.write_datetime(row_num+1, col_num, cell_format, value)
						else: ws.write(row_num+1, col_num, value, cell_format)
					else: ws.write_url(row_num+1, col_num, hLink, URL_center, string = value)
		print('OK')
		print('Write to file...', end=' ')
		while True:
			try:
				workbook.close()
				print('Ok')
				break
			except xlsxwriter.exceptions.FileCreateError as e:
				print(e)
				input()
		
	
	
if __name__ == '__main__':

	R = Letter()
	R.ReadXLS()
	R.SortFiles()
	R.UpdateLetters()
	R.WriteXLS()
	
		
	# There is no way to specify “AutoFit” for a column in the Excel file format. This feature is only
	# available at runtime from within Excel. It is possible to simulate “AutoFit” in your application by
	# tracking the maximum width of the data in the column as your write it and then adjusting the
	# column width at the end.
	